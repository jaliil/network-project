import json
import openpyxl
import subprocess
import platform
import logging
import time
import re
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.sessions.models import Session
from django.utils import timezone
from django.contrib.auth.models import User
from django.contrib.auth.views import LoginView
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q

from .models import Device, BTS, ActivityLog, Province, CommandHistory, UserProvinceCredential, ConnectionLog, NetworkLink
from .forms import DeviceForm

# اضافه شدن تسک جدید برای اجرای کلاینت‌های لایه 2 (CPE)
from .tasks import execute_network_commands, execute_cpe_mac_commands

from .utils.search import search_mac_in_network
from .utils.mikrotik import report_signal_strength, report_customers
from .utils.cisco import run_cisco_web, report_switch_web
from .utils.snmp_tools import get_snmp_traffic, get_device_interfaces

import routeros_api
from netmiko import ConnectHandler

logger = logging.getLogger(__name__)


def get_base_context():
    pending = list(CommandHistory.objects.filter(status='Pending').order_by('-executed_at'))
    return {
        'pending_tasks': pending,
        'notifications': pending, 
    }

class CustomLoginView(LoginView):
    template_name = 'registration/login.html'

@login_required(login_url='/login/')
def dashboard_view(request):
    if request.user.groups.filter(name='employee_user').exists():
        return redirect('employee_panel')

    sessions = Session.objects.filter(expire_date__gte=timezone.now())
    uid_list = [session.get_decoded().get('_auth_user_id') for session in sessions if session.get_decoded().get('_auth_user_id')]
    
    online_users = User.objects.filter(id__in=uid_list)
    total_users_count = User.objects.count()
    
    provinces = Province.objects.all()
    device_types = Device.DEVICE_TYPES
    
    chart_data = {}
    for prov in provinces:
        prov_data = {}
        for dev_val, dev_name in device_types:
            prov_data[dev_val] = Device.objects.filter(bts__province=prov, device_type=dev_val).count()
        chart_data[prov.name] = prov_data

    today_dt = timezone.localtime(timezone.now())
    today_date = today_dt.date()
    
    chart_labels = []
    mikrotik_trend = [0] * 30
    cisco_trend = [0] * 30
    
    for i in range(29, -1, -1):
        target_d = today_date - timedelta(days=i)
        chart_labels.append(target_d.strftime('%b %d'))
        
    start_dt = today_dt - timedelta(days=29)
    start_dt = start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    
    recent_tasks = CommandHistory.objects.filter(executed_at__gte=start_dt)
    
    for task in recent_tasks:
        if task.executed_at:
            try:
                t_date = timezone.localtime(task.executed_at).date()
            except:
                t_date = task.executed_at.date() if hasattr(task.executed_at, 'date') else None
                
            if t_date:
                day_diff = (today_date - t_date).days
                idx = 29 - day_diff
                if 0 <= idx <= 29:
                    if task.device_type == 'mikrotik':
                        mikrotik_trend[idx] += 1
                    elif task.device_type == 'cisco':
                        cisco_trend[idx] += 1
    
    usage_trend_data = json.dumps({
        'labels': chart_labels,
        'mikrotik': mikrotik_trend,
        'cisco': cisco_trend
    })

    pending_mikrotik = CommandHistory.objects.filter(device_type='mikrotik', status='Pending').count()
    done_mikrotik = CommandHistory.objects.filter(device_type='mikrotik', status='Completed').count()
    pending_cisco = CommandHistory.objects.filter(device_type='cisco', status='Pending').count()
    done_cisco = CommandHistory.objects.filter(device_type='cisco', status='Completed').count()

    context = get_base_context()
    context.update({
        'online_users': online_users, 
        'total_users_count': total_users_count,
        'provinces': provinces,
        'device_chart_data': json.dumps(chart_data), 
        'device_types': device_types, 
        'pending_mikrotik': pending_mikrotik,
        'done_mikrotik': done_mikrotik, 
        'pending_cisco': pending_cisco, 
        'done_cisco': done_cisco,
        'usage_trend_data': usage_trend_data,
    })
    return render(request, 'dashboard.html', context)

@login_required(login_url='/login/')
def run_mikrotik_view(request):
    searched_prov, entered_commands, entered_description = None, "", ""
    if request.method == 'POST':
        prov_id = request.POST.get('province')
        bts_ids = request.POST.getlist('bts')
        site_ips = request.POST.getlist('site') 
        description = request.POST.get('description', '')
        commands_text = request.POST.get('commands', '')
        entered_commands, entered_description = commands_text, description
        if prov_id and commands_text.strip():
            try:
                searched_prov = Province.objects.get(id=prov_id)
                if site_ips and any(ip.strip() for ip in site_ips):
                    valid_ips = [ip.strip() for ip in site_ips if ip.strip()]
                    devices = Device.objects.filter(ip_address__in=valid_ips, device_type='mikrotik')
                elif bts_ids and any(bts.strip() for bts in bts_ids):
                    valid_bts = [bts.strip() for bts in bts_ids if bts.strip()]
                    devices = Device.objects.filter(bts_id__in=valid_bts, device_type='mikrotik')
                else:
                    devices = Device.objects.filter(bts__province=searched_prov, device_type='mikrotik')
                ips_list = [dev.ip_address for dev in devices if dev.ip_address]
                if ips_list:
                    CommandHistory.objects.create(
                        user=request.user, description=description, commands=commands_text, 
                        device_type='mikrotik', target_ips=",".join(ips_list), 
                        status='Pending', total_devices=len(ips_list)
                    )
                    messages.success(request, f"Request for {len(ips_list)} devices submitted to DCN for approval.")
                else:
                    messages.warning(request, "No devices found matching these criteria.")
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")

    context = get_base_context()
    context.update({
        'provinces': Province.objects.all(), 'btss': BTS.objects.select_related('province').all(), 
        'histories': CommandHistory.objects.filter(device_type='mikrotik').order_by('-executed_at')[:15], 
        'searched_prov': searched_prov, 'entered_commands': entered_commands, 'entered_description': entered_description
    })
    return render(request, 'run_mikrotik.html', context)

@login_required(login_url='/login/')
def run_cisco_view(request):
    searched_prov, entered_commands, entered_description = None, "", ""
    if request.method == 'POST':
        prov_id = request.POST.get('province')
        bts_ids = request.POST.getlist('bts')
        site_ips = request.POST.getlist('site') 
        description = request.POST.get('description', '')
        commands_text = request.POST.get('commands', '')
        entered_commands, entered_description = commands_text, description
        if prov_id and commands_text.strip():
            try:
                searched_prov = Province.objects.get(id=prov_id)
                if site_ips and any(ip.strip() for ip in site_ips):
                    valid_ips = [ip.strip() for ip in site_ips if ip.strip()]
                    devices = Device.objects.filter(ip_address__in=valid_ips, device_type='cisco')
                elif bts_ids and any(bts.strip() for bts in bts_ids):
                    valid_bts = [bts.strip() for bts in bts_ids if bts.strip()]
                    devices = Device.objects.filter(bts_id__in=valid_bts, device_type='cisco')
                else:
                    devices = Device.objects.filter(bts__province=searched_prov, device_type='cisco')
                ips_list = [dev.ip_address for dev in devices if dev.ip_address]
                if ips_list:
                    CommandHistory.objects.create(
                        user=request.user, description=description, commands=commands_text, 
                        device_type='cisco', target_ips=",".join(ips_list), 
                        status='Pending', total_devices=len(ips_list)
                    )
                    messages.success(request, f"Request for {len(ips_list)} devices submitted to DCN for approval.")
                else:
                    messages.warning(request, "No devices found matching these criteria.")
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")

    context = get_base_context()
    context.update({
        'provinces': Province.objects.all(), 'btss': BTS.objects.select_related('province').all(), 
        'histories': CommandHistory.objects.filter(device_type='cisco').order_by('-executed_at')[:15], 
        'searched_prov': searched_prov, 'entered_commands': entered_commands, 'entered_description': entered_description
    })
    return render(request, 'run_cisco.html', context)


# ==========================================
# 🔴 بخش جدید: ویوی اجرای دستورات روی CPE مشتریان
# ==========================================
@login_required(login_url='/login/')
def customer_configs_view(request):
    searched_prov, entered_commands, entered_description = None, "", ""
    if request.method == 'POST':
        prov_id = request.POST.get('province')
        bts_id = request.POST.get('bts') # اینجا BTS همان QRT یا Sender است
        cpe_user = request.POST.get('cpe_user', '') # یوزر مشتریان
        cpe_pass = request.POST.get('cpe_pass', '') # رمز مشتریان
        description = request.POST.get('description', '')
        commands_text = request.POST.get('commands', '')
        
        entered_commands, entered_description = commands_text, description
        
        if prov_id and bts_id and commands_text.strip():
            try:
                # پیدا کردن دستگاه QRT (Sender)
                sender_device = Device.objects.filter(bts_id=bts_id, device_type='mikrotik').first()
                if sender_device and sender_device.ip_address:
                    
                    # ذخیره در دیتابیس برای اجرای بک‌گراند (با تگ مخصوص cpe_mac)
                    # یوزر و پسورد کلاینت‌ها را در description موقتا ذخیره می‌کنیم تا تسک سلری بتواند بخواند
                    CommandHistory.objects.create(
                        user=request.user, 
                        description=f"{description} | CPE_CREDS:{cpe_user}:{cpe_pass}", 
                        commands=commands_text, 
                        device_type='cpe_mac', # تگ مخصوص برای کلاینت‌های لایه ۲
                        target_ips=sender_device.ip_address, # ای‌پی QRT
                        status='Pending', 
                        total_devices=1 # در واقع 1 سندر است که داخلش چندین کلاینت پیدا می‌شود
                    )
                    messages.success(request, f"Request to run on ALL clients of {sender_device.ip_address} submitted for approval.")
                else:
                    messages.warning(request, "No Sender IP found for this BTS.")
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")

    context = get_base_context()
    context.update({
        'provinces': Province.objects.all(), 
        'btss': BTS.objects.select_related('province').all(), 
        'histories': CommandHistory.objects.filter(device_type='cpe_mac').order_by('-executed_at')[:15], 
        'entered_commands': entered_commands, 
        'entered_description': entered_description
    })
    return render(request, 'customer_configs.html', context)


# ==========================================
# 🔴 بخش آپدیت شده: اتصال Approve به موتور مک‌تلنت
# ==========================================
@login_required(login_url='/login/')
def approve_command_view(request, history_id):
    history = get_object_or_404(CommandHistory, id=history_id)
    if history.status == 'Pending':
        history.status = 'Running'
        history.save()
        
        ips_list = history.target_ips.split(',')
        dev = Device.objects.filter(ip_address=ips_list[0]).first()
        prov = dev.bts.province
        raw_commands = [cmd.strip() for cmd in history.commands.split('\n') if cmd.strip()]
        
        if history.device_type == 'mikrotik': 
            execute_network_commands.delay(history_id=history.id, ips_list=ips_list, username=prov.mt_user, password=prov.mt_pass, port=prov.mt_port, commands=raw_commands, device_type='mikrotik')
        
        elif history.device_type == 'cisco': 
            execute_network_commands.delay(history_id=history.id, ips_list=ips_list, username=prov.cisco_user, password=prov.cisco_pass, port=prov.cisco_port, commands=raw_commands, device_type='cisco')
        
        # اجرای موتور کلاینت‌های لایه 2
        elif history.device_type == 'cpe_mac':
            execute_cpe_mac_commands.delay(
                history_id=history.id, 
                sender_ip=ips_list[0], 
                sender_username=prov.mt_user, 
                sender_password=prov.mt_pass, 
                sender_port=prov.mt_port, 
                commands=raw_commands
            )
            
        messages.success(request, "Request approved and running in background.")
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))

@login_required(login_url='/login/')
def reject_command_view(request, history_id):
    history = get_object_or_404(CommandHistory, id=history_id)
    if history.status == 'Pending':
        history.status = 'Rejected'
        history.save()
        messages.success(request, "Request rejected successfully.")
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))

@login_required(login_url='/login/')
def add_device_view(request):
    if request.method == 'POST':
        form = DeviceForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Device saved successfully!')
            return redirect(request.path)
    else: 
        form = DeviceForm()
    context = get_base_context()
    context.update({
        'form': form, 'last_device': Device.objects.order_by('-id').first(), 
        'provinces': Province.objects.all(), 'bts_list': BTS.objects.select_related('province').all()
    })
    return render(request, 'add_device.html', context)

@login_required(login_url='/login/')
def bulk_upload_view(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            wb = openpyxl.load_workbook(request.FILES['excel_file'])
            ws = wb.active
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                bts_name, dev_type, ip, mac, mt_model, rl_model, ssid, freq = row
                if bts_name and dev_type and ip and mac:
                    bts_obj = BTS.objects.filter(name=bts_name).first()
                    if bts_obj:
                        Device.objects.create(bts=bts_obj, device_type=dev_type, ip_address=ip, mac_address=mac, device_model=mt_model or rl_model, ssid=ssid, frequency=freq)
                        count += 1
            messages.success(request, f"Success: {count} devices added!")
        except Exception as e: 
            messages.error(request, str(e))
        return redirect('bulk_upload')
    return render(request, 'bulk_upload.html', get_base_context())

@login_required(login_url='/login/')
def export_devices_view(request):
    if request.method == 'POST':
        prov_id, bts_id, dev_type = request.POST.get('province'), request.POST.get('bts'), request.POST.get('device_type')
        devices = Device.objects.all().select_related('bts', 'bts__province')
        if prov_id and prov_id != 'all': devices = devices.filter(bts__province_id=prov_id)
        if bts_id and bts_id != 'all': devices = devices.filter(bts_id=bts_id)
        if dev_type and dev_type != 'all': devices = devices.filter(device_type=dev_type)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['BTS Name', 'Province', 'Device Type', 'IP Address', 'MAC Address', 'Device Model', 'SSID', 'Frequency'])
        for d in devices: ws.append([d.bts.name, d.bts.province.name, d.get_device_type_display(), d.ip_address, d.mac_address, d.get_device_model_display(), d.ssid, d.frequency])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Export.xlsx"'
        wb.save(response)
        return response
    context = get_base_context()
    context.update({'provinces': Province.objects.all(), 'btss': BTS.objects.all(), 'device_types': Device.DEVICE_TYPES})
    return render(request, 'export_filter.html', context)

@login_required(login_url='/login/')
def device_list(request):
    devices = Device.objects.select_related('bts', 'bts__province').all().order_by('-id')
    q, prov, dev = request.GET.get('search', ''), request.GET.get('province', ''), request.GET.get('device_type', '')
    if q: devices = devices.filter(Q(ip_address__icontains=q)|Q(mac_address__icontains=q)|Q(bts__name__icontains=q))
    if prov: devices = devices.filter(bts__province_id=prov)
    if dev: devices = devices.filter(device_type=dev)
    paginator = Paginator(devices, 20)
    try: page_obj = paginator.page(request.GET.get('page'))
    except: page_obj = paginator.page(1)
    context = get_base_context()
    context.update({'devices': page_obj, 'page_obj': page_obj, 'provinces': Province.objects.all()})
    return render(request, 'device_list.html', context)

@login_required(login_url='/login/')
def search_mac_view(request):
    result = None
    if request.method == 'POST':
        mac_query, prov_id, dev_type = request.POST.get('mac_address'), request.POST.get('province'), request.POST.get('device_type')
        if mac_query and prov_id and dev_type:
            try:
                prov = Province.objects.get(id=prov_id)
                devices = Device.objects.filter(bts__province=prov, device_type=dev_type)
                host_list = [dev.ip_address for dev in devices if dev.ip_address]
                if host_list:
                    u, p, po = (prov.mt_user, prov.mt_pass, prov.mt_port) if dev_type == 'mikrotik' else (prov.cisco_user, prov.cisco_pass, prov.cisco_port)
                    result = search_mac_in_network(host_list, dev_type, mac_query, u, p, po)
                else: result = {"status": "error", "message": "No devices found."}
            except Exception as e: result = {"status": "error", "message": str(e)}
    context = get_base_context()
    context.update({'provinces': Province.objects.all(), 'result': result})
    return render(request, 'search_mac.html', context)

@login_required(login_url='/login/')
def check_signal_view(request):
    results, searched_prov = None, None
    if request.method == 'POST':
        prov_id, bts_ids, site_ips = request.POST.get('province'), request.POST.getlist('bts'), request.POST.getlist('site')
        try: threshold = int(request.POST.get('threshold', '-75'))
        except: threshold = -75
        if prov_id:
            try:
                searched_prov = Province.objects.get(id=prov_id)
                if site_ips and any(ip.strip() for ip in site_ips):
                    valid_ips = [ip.strip() for ip in site_ips if ip.strip()]
                    devices = Device.objects.filter(ip_address__in=valid_ips, device_type='mikrotik')
                elif bts_ids and any(bts.strip() for bts in bts_ids):
                    valid_bts = [bts.strip() for bts in bts_ids if bts.strip()]
                    devices = Device.objects.filter(bts_id__in=valid_bts, device_type='mikrotik')
                else:
                    devices = Device.objects.filter(bts__province=searched_prov, device_type='mikrotik')

                results = []
                for dev in devices:
                    if dev.ip_address:
                        res = report_signal_strength(dev.ip_address, searched_prov.mt_user, searched_prov.mt_pass, searched_prov.mt_port, threshold)
                        res['bts_name'] = dev.bts.name
                        results.append(res)
            except Exception as e: results = [{"status": "error", "error": str(e)}]
    context = get_base_context()
    context.update({'provinces': Province.objects.all(), 'btss': BTS.objects.select_related('province').all(), 'results': results, 'searched_prov': searched_prov})
    return render(request, 'check_signal.html', context)

@login_required(login_url='/login/')
def wireless_clients_view(request):
    results, searched_prov = None, None
    if request.method == 'POST':
        prov_id, bts_ids, site_ips = request.POST.get('province'), request.POST.getlist('bts'), request.POST.getlist('site')
        if prov_id:
            try:
                searched_prov = Province.objects.get(id=prov_id)
                if site_ips and any(ip.strip() for ip in site_ips):
                    valid_ips = [ip.strip() for ip in site_ips if ip.strip()]
                    devices = Device.objects.filter(ip_address__in=valid_ips, device_type='mikrotik')
                elif bts_ids and any(bts.strip() for bts in bts_ids):
                    valid_bts = [bts.strip() for bts in bts_ids if bts.strip()]
                    devices = Device.objects.filter(bts_id__in=valid_bts, device_type='mikrotik')
                else:
                    devices = Device.objects.filter(bts__province=searched_prov, device_type='mikrotik')
                
                results = []
                for dev in devices:
                    if dev.ip_address:
                        res = report_customers(dev.ip_address, searched_prov.mt_user, searched_prov.mt_pass, searched_prov.mt_port)
                        res['bts_name'], res['ssid'] = dev.bts.name, dev.ssid
                        results.append(res)
            except Exception as e: results = [{"status": "error", "error": str(e)}]
    context = get_base_context()
    context.update({'provinces': Province.objects.all(), 'btss': BTS.objects.select_related('province').all(), 'sites': Device.objects.filter(device_type='mikrotik'), 'results': results, 'searched_prov': searched_prov})
    return render(request, 'wireless_clients.html', context)

@login_required(login_url='/login/')
def switch_report_view(request):
    results, searched_prov = None, None
    if request.method == 'POST':
        prov_id, bts_ids, site_ips = request.POST.get('province'), request.POST.getlist('bts'), request.POST.getlist('site')
        if prov_id:
            try:
                searched_prov = Province.objects.get(id=prov_id)
                if site_ips and any(ip.strip() for ip in site_ips):
                    valid_ips = [ip.strip() for ip in site_ips if ip.strip()]
                    devices = Device.objects.filter(ip_address__in=valid_ips, device_type='cisco')
                elif bts_ids and any(bts.strip() for bts in bts_ids):
                    valid_bts = [bts.strip() for bts in bts_ids if bts.strip()]
                    devices = Device.objects.filter(bts_id__in=valid_bts, device_type='cisco')
                else:
                    devices = Device.objects.filter(bts__province=searched_prov, device_type='cisco')
                
                results = []
                for dev in devices:
                    if dev.ip_address:
                        res = report_switch_web(dev.ip_address, searched_prov.cisco_user, searched_prov.cisco_pass, searched_prov.cisco_port)
                        res['bts_name'] = dev.bts.name
                        results.append(res)
            except Exception as e: results = [{"status": "error", "error": str(e)}]
    context = get_base_context()
    context.update({'provinces': Province.objects.all(), 'btss': BTS.objects.select_related('province').all(), 'sites': Device.objects.filter(device_type='cisco'), 'results': results, 'searched_prov': searched_prov})
    return render(request, 'switch_report.html', context)

@login_required(login_url='/login/')
def employee_panel_view(request):
    user_creds = UserProvinceCredential.objects.filter(user=request.user)
    creds_dict = {}
    for c in user_creds:
        creds_dict[c.province.id] = {
            's_pass': c.sender_pass or '',
            'r_pass': c.receiver_pass or ''
        }

    context = get_base_context()
    context.update({
        'target_ip': request.GET.get('target_ip'),
        'provinces': Province.objects.all(),
        'user_creds_json': json.dumps(creds_dict)
    })
    return render(request, 'employee_panel.html', context)

@login_required(login_url='/login/')
def update_profile(request):
    provinces = Province.objects.all()
    if request.method == 'POST':
        prov_id = request.POST.get('province_id')
        if prov_id:
            province = get_object_or_404(Province, id=prov_id)
            s_pass = request.POST.get('sender_password')
            r_pass = request.POST.get('receiver_password')
            cred, created = UserProvinceCredential.objects.get_or_create(user=request.user, province=province)
            cred.sender_pass = s_pass
            cred.receiver_pass = r_pass
            cred.save()
            messages.success(request, f"Credentials for {province.name} updated successfully.")
            return redirect('update_profile')
            
    user_creds = UserProvinceCredential.objects.filter(user=request.user)
    creds_dict = {}
    for c in user_creds:
        creds_dict[c.province.id] = {'s_pass': c.sender_pass or '', 'r_pass': c.receiver_pass or ''}

    return render(request, 'update_profile.html', {
        'provinces': provinces, 'creds_json': json.dumps(creds_dict)
    })

@login_required(login_url='/login/')
def smart_ping_and_log(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            ip_address = data.get('ip_address')
            province_id = data.get('province_id')
            device_type = data.get('device_type', 'unknown')

            if not ip_address:
                return JsonResponse({'status': 'error', 'message': 'IP address is required.'})

            param = '-n' if platform.system().lower() == 'windows' else '-c'
            command = ['ping', param, '1', '-W', '1', ip_address]
            
            try:
                output = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                is_online = (output.returncode == 0)
            except Exception:
                is_online = False

            province = Province.objects.filter(id=province_id).first() if province_id else None
            ConnectionLog.objects.create(
                user=request.user, province=province, ip_address=ip_address,
                device_type=device_type, is_online=is_online
            )
            return JsonResponse({'status': 'success', 'is_online': is_online})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})

@login_required(login_url='/login/')
def live_map_view(request):
    provinces = Province.objects.all().order_by('name')
    bts_list = BTS.objects.exclude(latitude__isnull=True).exclude(longitude__isnull=True)
    
    context = get_base_context()
    context.update({ 
        'bts_list': bts_list, 
        'provinces': provinces 
    })
    return render(request, 'live_map.html', context)

@login_required(login_url='/login/')
def map_management_view(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add_link':
            source_id = request.POST.get('source_bts')
            target_id = request.POST.get('target_bts')
            link_type = request.POST.get('link_type')
            capacity = request.POST.get('capacity_mbps', 1000)
            
            src_device = request.POST.get('source_device_type', 'mikrotik')
            src_ip = request.POST.get('source_ip')
            src_interface = request.POST.get('source_interface')
            snmp_community = request.POST.get('snmp_community', 'public')
            
            tgt_device = request.POST.get('target_device_type', 'mikrotik')
            tgt_ip = request.POST.get('target_ip')
            tgt_interface = request.POST.get('target_interface')
            tgt_snmp_community = request.POST.get('target_snmp_community', 'public')

            if source_id == target_id:
                messages.error(request, "Error: Source and Target BTS cannot be the same!")
            elif not source_id or not target_id:
                messages.error(request, "Please select both Source and Target BTS.")
            else:
                try:
                    NetworkLink.objects.create(
                        source_bts_id=source_id,
                        target_bts_id=target_id,
                        source_device_type=src_device,
                        source_ip=src_ip,
                        source_interface=src_interface,
                        snmp_community=snmp_community,
                        target_device_type=tgt_device,
                        target_ip=tgt_ip,
                        target_interface=tgt_interface,
                        target_snmp_community=tgt_snmp_community,
                        link_type=link_type,
                        capacity_mbps=capacity
                    )
                    messages.success(request, "Network link successfully created!")
                except Exception as e:
                    messages.error(request, f"Database Error: {str(e)}")
        
        elif action == 'edit_link':
            link_id = request.POST.get('link_id')
            try:
                link = NetworkLink.objects.get(id=link_id)
                link.link_type = request.POST.get('link_type')
                link.capacity_mbps = request.POST.get('capacity_mbps')
                
                if 'source_device_type' in request.POST:
                    link.source_device_type = request.POST.get('source_device_type')
                if 'target_device_type' in request.POST:
                    link.target_device_type = request.POST.get('target_device_type')
                if 'snmp_community' in request.POST:
                    link.snmp_community = request.POST.get('snmp_community')
                if 'target_snmp_community' in request.POST:
                    link.target_snmp_community = request.POST.get('target_snmp_community')
                    
                link.save()
                messages.success(request, "Link updated successfully.")
            except Exception as e:
                messages.error(request, f"Error updating link: {str(e)}")

        elif action == 'delete_link':
            link_id = request.POST.get('link_id')
            try:
                NetworkLink.objects.get(id=link_id).delete()
                messages.success(request, "Link deleted successfully.")
            except:
                messages.error(request, "Error deleting the link.")
                
        return redirect('map_management')

    context = get_base_context()
    context.update({
        'provinces': Province.objects.all().order_by('name'), 
        'btss': BTS.objects.select_related('province').all().order_by('province__name', 'name'),
        'links': NetworkLink.objects.select_related('source_bts', 'target_bts').all().order_by('-id'),
        'link_types': NetworkLink.LINK_TYPES
    })
    return render(request, 'map_management.html', context)

@login_required(login_url='/login/')
def fetch_interfaces_view(request):
    ip_address = request.GET.get('ip')
    community = request.GET.get('community', 'public') 
    
    if not ip_address:
        return JsonResponse({'status': 'error', 'message': 'IP Address missing'})
    try:
        interfaces = get_device_interfaces(ip_address, community)
        
        if interfaces:
            return JsonResponse({'status': 'success', 'interfaces': interfaces})
        else:
            return JsonResponse({'status': 'error', 'message': 'No interfaces found or SNMP timeout.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@login_required(login_url='/login/')
def live_map_data_api(request):
    links = NetworkLink.objects.filter(is_active=True).select_related('source_bts', 'target_bts')
    links_data = []
    now = timezone.now()
    
    for link in links:
        total_traffic = link.current_rx_mbps + link.current_tx_mbps
        usage_percent = 0
        if link.capacity_mbps > 0:
            usage_percent = (total_traffic / link.capacity_mbps) * 100

        is_down = False
        if not link.last_snmp_update:
            is_down = True
        else:
            time_difference = now - link.last_snmp_update
            if time_difference.total_seconds() > 300:
                is_down = True
                
        link_status = 'down' if is_down else 'up'

        links_data.append({
            'link_id': link.id,
            'source_id': link.source_bts.id,
            'target_id': link.target_bts.id,
            'tx_mbps': link.current_tx_mbps,
            'rx_mbps': link.current_rx_mbps,
            'capacity': link.capacity_mbps,
            'usage_percent': round(usage_percent, 1),
            'type': link.link_type,
            'status': link_status,
            'is_active': link.is_active
        })
        
    return JsonResponse({'status': 'success', 'links': links_data})


# ==========================================
# 🔴 سیستم هوشمند تحلیل لاگ (Smart Log Analyzer) - با netmiko و routeros_api
# ==========================================

def parse_device_time(time_str):
    """
    تابع هوشمند برای تبدیل زمان‌های عجیب میکروتیک و سیسکو به زمان استاندارد پایتون جهت مقایسه
    """
    if not time_str or time_str == 'unknown':
        return None
    time_str = time_str.strip('* ')
    now = timezone.now()
    try:
        # حالت 1: فقط ساعت (مثلاً 06:37:20) - در میکروتیک یعنی امروز
        if re.match(r'^\d{2}:\d{2}:\d{2}$', time_str):
            t = datetime.strptime(time_str, '%H:%M:%S').time()
            return datetime.combine(now.date(), t)
        
        # حالت 2: تاریخ میکروتیک (مثلاً aug/21 06:37:20)
        if re.match(r'^[a-zA-Z]{3}/\d{2}\s\d{2}:\d{2}:\d{2}$', time_str):
            dt = datetime.strptime(time_str, '%b/%d %H:%M:%S')
            return dt.replace(year=now.year)
            
        # حالت 3: تاریخ سیسکو (مثلاً Aug 21 06:37:20)
        if re.match(r'^[a-zA-Z]{3}\s+\d{1,2}\s\d{2}:\d{2}:\d{2}', time_str):
            clean_str = time_str.split('.')[0]
            clean_str = re.sub(r'\s+', ' ', clean_str) # حذف فاصله‌های اضافی
            dt = datetime.strptime(clean_str, '%b %d %H:%M:%S')
            return dt.replace(year=now.year)
            
        # حالت 4: تاریخ استاندارد
        if re.match(r'^\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}:\d{2}', time_str):
            clean_str = time_str.split('.')[0]
            return datetime.strptime(clean_str, '%Y-%m-%d %H:%M:%S')
    except:
        pass
    return None

@login_required(login_url='/login/')
def device_logs_analyzer_view(request):
    """
    نمایش رابط کاربری ترمینال و ارسال اطلاعات استان‌ها، دکل‌ها و نوع دستگاه‌ها از دیتابیس
    """
    context = get_base_context()
    context.update({
        'provinces': Province.objects.all(),
        'btss': BTS.objects.select_related('province').all(),
        'device_types': Device.DEVICE_TYPES,
    })
    return render(request, 'device_logs.html', context)


@login_required(login_url='/login/')
def fetch_device_logs_api(request):
    """
    اتصال زنده به دستگاه (از طریق Netmiko و RouterOS API) بر اساس IP دستی و Credentials استان
    """
    if request.method == 'POST':
        try:
            province_id = request.POST.get('province_id')
            device_type = request.POST.get('device_type')
            device_ip = request.POST.get('device_ip')
            severities_json = request.POST.get('severities', '[]')
            severities = json.loads(severities_json)

            # دریافت زمان شروع و پایان از رابط کاربری
            start_time_str = request.POST.get('start_time')
            end_time_str = request.POST.get('end_time')
            start_dt, end_dt = None, None
            
            try:
                if start_time_str: start_dt = datetime.strptime(start_time_str, '%Y-%m-%dT%H:%M')
                if end_time_str: end_dt = datetime.strptime(end_time_str, '%Y-%m-%dT%H:%M')
            except Exception:
                pass

            if not device_ip or not province_id or not device_type or not severities:
                return JsonResponse({'status': 'error', 'message': 'Missing required parameters (Province, Type, IP, or Severity).'})

            prov = Province.objects.filter(id=province_id).first()
            if not prov:
                return JsonResponse({'status': 'error', 'message': 'Selected Province not found.'})

            username = prov.mt_user if device_type == 'mikrotik' else prov.cisco_user
            password = prov.mt_pass if device_type == 'mikrotik' else prov.cisco_pass
            port = prov.mt_port if device_type == 'mikrotik' else prov.cisco_port

            logs_output = []

            # ==============================
            # منطق میکروتیک (RouterOS API)
            # ==============================
            if device_type == 'mikrotik':
                connection = None
                try:
                    connection = routeros_api.RouterOsApiPool(
                        host=device_ip, username=username, password=password,
                        port=port, plaintext_login=True
                    )
                    api = connection.get_api()
                    log_resource = api.get_resource('/log')
                    
                    raw_logs = log_resource.get()
                    
                    for log in raw_logs[-300:]: # بررسی 300 لاگ آخر
                        topics = log.get('topics', '').lower()
                        time_str = log.get('time', 'unknown')
                        msg = log.get('message', '')

                        # فیلتر زمان
                        if start_dt or end_dt:
                            log_dt = parse_device_time(time_str)
                            if log_dt:
                                if start_dt and log_dt < start_dt: continue
                                if end_dt and log_dt > end_dt: continue

                        sev_type = 'info'
                        
                        # اضافه شدن Alert به فیلتر میکروتیک
                        if 'alert' in severities and 'alert' in topics:
                            sev_type = 'alert'
                        elif 'critical' in severities and 'critical' in topics:
                            sev_type = 'critical'
                        elif 'error' in severities and 'error' in topics:
                            sev_type = 'error'
                        elif 'warning' in severities and 'warning' in topics:
                            sev_type = 'warning'

                        if sev_type != 'info':
                            logs_output.append({
                                'time': time_str,
                                'severity': sev_type,
                                'message': f"[{topics.upper()}] {msg}"
                            })
                except Exception as err:
                    return JsonResponse({'status': 'error', 'message': f'MikroTik Connection Error: {str(err)}'})
                finally:
                    if connection:
                        try:
                            connection.disconnect()
                        except:
                            pass

            # ==============================
            # منطق سیسکو (Netmiko)
            # ==============================
            elif device_type == 'cisco':
                net_connect = None
                try:
                    cisco_device = {
                        "device_type": "cisco_ios",
                        "ip": device_ip,
                        "username": username,
                        "password": password,
                        "port": port,
                    }
                    net_connect = ConnectHandler(**cisco_device)
                    output = net_connect.send_command("show logging")
                    
                    for line in output.splitlines():
                        if not line.strip():
                            continue
                            
                        # استخراج حدودی زمان از متن سیسکو
                        time_str = timezone.now().strftime('%H:%M:%S') 
                        match = re.search(r'^[A-Z][a-z]{2}\s+\d+\s+\d{2}:\d{2}:\d{2}', line.strip('* '))
                        if match: time_str = match.group()

                        # فیلتر زمان
                        if start_dt or end_dt:
                            log_dt = parse_device_time(time_str)
                            if log_dt:
                                if start_dt and log_dt < start_dt: continue
                                if end_dt and log_dt > end_dt: continue

                        sev_type = 'info'

                        # اضافه شدن Alert (کد -1-) به سیسکو
                        if 'alert' in severities and '-1-' in line:
                            sev_type = 'alert'
                        elif 'critical' in severities and ('-0-' in line or '-2-' in line):
                            sev_type = 'critical'
                        elif 'error' in severities and '-3-' in line:
                            sev_type = 'error'
                        elif 'warning' in severities and '-4-' in line:
                            sev_type = 'warning'

                        if sev_type != 'info':
                            logs_output.append({
                                'time': time_str,
                                'severity': sev_type,
                                'message': line.strip()
                            })
                except Exception as err:
                    return JsonResponse({'status': 'error', 'message': f'Cisco Connection Error: {str(err)}'})
                finally:
                    if net_connect:
                        try:
                            net_connect.disconnect()
                        except:
                            pass

            return JsonResponse({'status': 'success', 'logs': logs_output})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})


@login_required(login_url='/login/')
def analyze_log_ai_api(request):
    """
    سیستم تحلیل هوش مصنوعی (در حال حاضر غیرفعال - رزرو شده برای آینده)
    """
    if request.method == 'POST':
        # ایجاد یک توقف کوتاه (۱ ثانیه) تا کاربر افکت لودینگ روی دکمه را ببیند و حس طبیعی بودن بدهد
        time.sleep(1)
        
        # پیامی که در پنجره پاپ‌آپ به کاربر نشان داده می‌شود
        coming_soon_message = (
            "🚧 <b>AI Root Cause Analysis is currently disabled.</b><br><br>"
            "This feature is successfully integrated and is reserved for future activation. "
            "Once enabled, it will provide instant AI-driven troubleshooting for network logs."
        )
        
        return JsonResponse({'status': 'success', 'analysis': coming_soon_message})
            
    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})